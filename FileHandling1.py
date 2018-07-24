from openpyxl import workbook, Workbook


class startExecution(workbook):


    def __init__(self):
        pass



    def createworkbook(self):
        book = Workbook()
        sheet = book.active



    def readfile(self):
        book = Workbook()
        sheet = book.active

    def writefile(self):
        book = Workbook()
        sheet = book.active



def main():
    book = Workbook()
    name = startExecution()
    name.writefile()
    rows = (
        ('SR.', 'UserName', 'Passward', 'Expected-Msg'),
        (1, 'admin', 'admin123', 'Success'),
        (2, 'admin123', 'admin', 'Invalid User name or Passaward'),
        (3, 'NA', 'admin123', 'Username field cant be empty'),
        (4, 'admin', 'NA', 'Passward field cant be empty'),
        (5, 'abcpqr', 'xyz', 'Invalid')
    )

    for row in rows:
        name.sheet.append(row)

    book.save("Email.xlsx")


if __name__ == '__main__':
    main()

from openpyxl import Workbook
import openpyxl
import statistics as stats
'''
book = Workbook()
sheet = book.active

sheet['A1'] = 1
sheet.cell(row=5, column=5).value = 2

book.save('write2cell.xlsx')

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('appending.xlsx')



book = openpyxl.load_workbook('sample.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value)
print(a3.value)



book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

#for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
 #   for cell in row:
  #      print(cell.value, end=" ")
  #  print()

for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()

book.save('iterbycols.xlsx')
#book.save('iterbyrows.xlsx')




book = openpyxl.load_workbook('numbers.xlsx', data_only=True)

sheet = book.active

rows = sheet.rows

values = []

for row in rows:
    for cell in row:
        values.append(cell.value)

print("Number of values: {0}".format(len(values)))
print("Sum of values: {0}".format(sum(values)))
print("Minimum value: {0}".format(min(values)))
print("Maximum value: {0}".format(max(values)))
print("Mean: {0}".format(stats.mean(values)))
print("Median: {0}".format(stats.median(values)))
print("Standard deviation: {0}".format(stats.stdev(values)))
print("Variance: {0}".format(stats.variance(values)))
'''
from openpyxl import Workbook
from openpyxl.drawing.image import Image

book = Workbook()
sheet = book.active

img = Image("icesid.png")
sheet['A1'] = 'This is Sid'

sheet.add_image(img, 'B2')

book.save("sheet_image.xlsx")
